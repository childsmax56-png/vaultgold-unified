#!/usr/bin/env python3
"""Extract the real hyperlink URLs embedded in a Google Sheet's cells.

The public CSV / gviz export of a sheet only carries the DISPLAY TEXT of each
hyperlink (e.g. "Pixeldrain", "Download") — never the underlying href. Our
earlier extractors (scripts/extract-*.gs) worked around that with Google Apps
Script. This is the pure-Python equivalent: the sheet's XLSX export *does*
preserve embedded cell hyperlinks, so we download that and read the hrefs with
openpyxl. No OAuth and no edit access needed — anonymous "anyone with the link"
view access is enough.

Output matches the .gs extractors: a CSV with columns  Tab, Era, Name, URL.

Usage:
    python3 scripts/extract-sheet-links.py                 # default sheet below
    python3 scripts/extract-sheet-links.py --id <SHEET_ID>
    python3 scripts/extract-sheet-links.py --url "<share url>"
    python3 scripts/extract-sheet-links.py -o ~/Downloads/links.csv

Requires: requests, openpyxl   ->   pip install requests openpyxl
"""
import argparse
import csv
import re
import sys
import tempfile

# The sheet this was last pointed at. Override with --id / --url.
DEFAULT_SHEET_ID = "1a8_li_D3rG0iDLqT9AGZsRVojlEyhO_nb735cRpyUvE"

EXPORT_URL = "https://docs.google.com/spreadsheets/d/{id}/export?format=xlsx"

# Pull a sheet id out of a full share URL (.../spreadsheets/d/<ID>/edit...).
ID_IN_URL = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
URL_IN_TEXT = re.compile(r"https?://\S+")
HYPERLINK_FN = re.compile(r'HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE)


def resolve_id(args):
    if args.url:
        m = ID_IN_URL.search(args.url)
        if not m:
            sys.exit(f"Could not find a sheet id in URL: {args.url}")
        return m.group(1)
    return args.id


def download_xlsx(sheet_id):
    try:
        import requests
    except ImportError:
        sys.exit("Missing dependency: pip install requests openpyxl")
    url = EXPORT_URL.format(id=sheet_id)
    print(f"Downloading XLSX export of {sheet_id} ...")
    r = requests.get(url, timeout=60, allow_redirects=True)
    if r.status_code != 200:
        sys.exit(
            f"Export request failed (HTTP {r.status_code}). Make sure the sheet "
            f"is shared as 'anyone with the link can view'.\n  {url}"
        )
    ctype = r.headers.get("content-type", "")
    if "spreadsheet" not in ctype and "openxml" not in ctype:
        # Google serves an HTML sign-in / error page when the sheet is private.
        sys.exit(
            "The export did not return a spreadsheet — the sheet is probably "
            "not publicly viewable. Set sharing to 'anyone with the link'."
        )
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(r.content)
    tmp.close()
    print(f"  {len(r.content):,} bytes")
    return tmp.name


def clean_header(v):
    return str(v or "").replace("\n", " ").strip().lower()


def find_header_row(ws, scan=15):
    """Locate the header row and the Era / Name column indexes (0-based).

    Scans the first `scan` rows for one that names both an era- and a name-like
    column, mirroring the .gs extractors. Falls back to row 1 with no matched
    columns so link extraction still runs on unusual layouts.
    """
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=scan), start=1):
        era_idx = name_idx = -1
        for cidx, cell in enumerate(row):
            h = clean_header(cell.value)
            if era_idx < 0 and h.startswith("era"):
                era_idx = cidx
            if name_idx < 0 and (h.startswith("name") or h.startswith("title")):
                name_idx = cidx
        if era_idx >= 0 and name_idx >= 0:
            return ridx, era_idx, name_idx
    return 1, -1, -1


def urls_in_cell(cell):
    """Every URL reachable from one cell: embedded hyperlink, =HYPERLINK()
    formula, and any raw http(s) text — deduped, order preserved."""
    found = []

    def add(u):
        if u:
            u = u.strip().rstrip(",")
            if u.startswith("http") and u not in found:
                found.append(u)

    if cell.hyperlink and cell.hyperlink.target:
        add(cell.hyperlink.target)
    val = cell.value
    if isinstance(val, str):
        for m in HYPERLINK_FN.finditer(val):
            add(m.group(1))
        for m in URL_IN_TEXT.finditer(val):
            add(m.group(0))
    return found


def first_line(v):
    return str(v or "").split("\n")[0].strip()


def extract(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Missing dependency: pip install requests openpyxl")
    wb = load_workbook(path, data_only=False, read_only=False)
    out = [["Tab", "Era", "Name", "URL"]]
    for ws in wb.worksheets:
        hdr_row, era_idx, name_idx = find_header_row(ws)
        tab_total = 0
        for row in ws.iter_rows(min_row=hdr_row + 1):
            era = str(row[era_idx].value or "").strip() if era_idx >= 0 else ""
            name = first_line(row[name_idx].value) if name_idx >= 0 else ""
            # Skip count / section-summary rows (multi-line era cell), matching .gs.
            if era_idx >= 0 and "\n" in str(row[era_idx].value or ""):
                continue
            row_urls = []
            for cidx, cell in enumerate(row):
                if cidx in (era_idx, name_idx):
                    continue
                for u in urls_in_cell(cell):
                    if u not in row_urls:
                        row_urls.append(u)
            for u in row_urls:
                out.append([ws.title, era, name, u])
                tab_total += 1
        print(f"  {ws.title}: {tab_total} links")
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--id", default=DEFAULT_SHEET_ID, help="Google Sheet id")
    ap.add_argument("--url", help="Full share URL (id is parsed out of it)")
    ap.add_argument("-o", "--out", default="sheet-links.csv", help="Output CSV path")
    args = ap.parse_args()

    sheet_id = resolve_id(args)
    xlsx = download_xlsx(sheet_id)
    rows = extract(xlsx)

    with open(args.out, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    print(f"\nDone! {len(rows) - 1} links -> {args.out}")


if __name__ == "__main__":
    main()
